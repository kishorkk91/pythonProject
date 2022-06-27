import openpyxl

inv_file = openpyxl.load_workbook("working_with_spreadsheet/inventory.xlsx")
product_list = inv_file["Sheet1"]

# dictionary variable to store supplier name and product
products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

# print(product_list.max_row)

# Iterate on for loop from second row because first row is just headline and does not contain any product information
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # print(supplier_name)

    # Calculation number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        # Following way we can also get value of dictionary instead of products_per_supplier[supplier_name]
        # current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        print("Adding a new supplier")
        products_per_supplier[supplier_name] = 1

        # print(products_per_supplier)

    # Calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    #print(products_per_supplier)
    #print(total_value_per_supplier)

    # Calculation of products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    #print(products_under_10_inv)

    # Add value for total inventory price
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

inv_file.save("working_with_spreadsheet/inventory_with_total_value.xlsx")
